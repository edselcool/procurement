from flask import Flask, request, redirect, url_for, jsonify, abort, render_template, flash
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from datetime import datetime
import os
from flask import send_file
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill


app = Flask(__name__)

# ✅ Use DATABASE_URL if provided (PostgreSQL recommended), else fallback to SQLite
db_url = os.environ.get('DATABASE_URL', 'sqlite:///pms.db')

# Fix old-style postgres:// URLs for SQLAlchemy compatibility
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url

# ✅ Secret key (environment first, local fallback)
app.config['SECRET_KEY'] = os.environ.get('PMS_SECRET', 'dev-secret')

# ✅ Disable debug in production (Render will set this automatically)
DEBUG = os.environ.get('FLASK_DEBUG', '0') == '1'

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'


# ------------------ Models ------------------
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(30), default='requester')  # requester, approver, admin

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    contact = db.Column(db.String(200))
    email = db.Column(db.String(200))

class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    unit_price = db.Column(db.Float, default=0.0)

class PurchaseRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    status = db.Column(db.String(30), default='draft')  # draft, pending, approved, rejected
    total = db.Column(db.Float, default=0.0)
    line_items = db.relationship('LineItem', backref='purchase_request', lazy=True)
    purchase_orders = db.relationship('PurchaseOrder', backref='purchase_request', lazy=True)

class LineItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pr_id = db.Column(db.Integer, db.ForeignKey('purchase_request.id'))
    item_name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Integer, default=1)
    unit = db.Column(db.String(50), default='pcs')
    unit_price = db.Column(db.Float, default=0.0)


    @property
    def subtotal(self):
        return (self.quantity or 0) * (self.unit_price or 0)

class ApprovalLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pr_id = db.Column(db.Integer, db.ForeignKey('purchase_request.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(20), nullable=False)  # "approve" or "reject"
    comment = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class PurchaseOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pr_id = db.Column(db.Integer, db.ForeignKey('purchase_request.id'), nullable=False)
    item_id = db.Column(db.Integer, db.ForeignKey('line_item.id'), nullable=False)
    supplier_name = db.Column(db.String(120), nullable=False)
    quotation_price = db.Column(db.Float, nullable=False)
    brand_name = db.Column(db.String(120), nullable=True)  # You requested this
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    item = db.relationship('LineItem', backref='purchase_orders')

class Balance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    activity_name = db.Column(db.String(255), nullable=False)  # same as PR.title
    pr_total_amount = db.Column(db.Float, default=0.0)
    po_total_amount = db.Column(db.Float, default=0.0)
    balance_amount = db.Column(db.Float, default=0.0)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    pr_id = db.Column(db.Integer, db.ForeignKey('purchase_request.id'), nullable=False)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

def compute_pr_total(pr_id):
    # sum of PR line items (use pr.total if you already maintain it)
    pr = PurchaseRequest.query.get(pr_id)
    if not pr:
        return 0.0
    # prefer stored pr.total if available
    try:
        return float(pr.total or 0.0)
    except:
        # fallback to summing line items
        return sum((li.quantity or 0) * (li.unit_price or 0) for li in LineItem.query.filter_by(pr_id=pr_id).all())

def compute_po_total_for_pr(pr_id):
    # sum of quotation_price for all POs that reference this PR
    total = db.session.query(db.func.coalesce(db.func.sum(PurchaseOrder.quotation_price), 0.0))\
        .filter(PurchaseOrder.pr_id == pr_id).scalar()
    return float(total or 0.0)

def update_balance_for_pr(pr_id):
    pr = PurchaseRequest.query.get(pr_id)
    if not pr:
        return

    # Total PR amount (sum of all line item subtotals)
    pr_total = sum([li.subtotal for li in pr.line_items])

    # Total PO amount (sum of quotation_price * quantity)
    po_total = sum([
        po.quotation_price * (po.item.quantity if po.item else 0)
        for po in pr.purchase_orders
    ])

    balance = pr_total - po_total

    bal = Balance.query.filter_by(pr_id=pr_id).first()
    if not bal:
        bal = Balance(
            activity_name=pr.title,
            pr_total_amount=pr_total,
            po_total_amount=po_total,
            balance_amount=balance,
            user_id=pr.created_by,
            pr_id=pr.id
        )
        db.session.add(bal)
    else:
        bal.pr_total_amount = pr_total
        bal.po_total_amount = po_total
        bal.balance_amount = balance

    db.session.commit()

def recalc_all_balances():
    prs = PurchaseRequest.query.all()
    for pr in prs:
        update_balance_for_pr(pr.id)

# ------------------ Helpers ------------------
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def role_required(*roles):
    """Decorator to require one of the roles (strings) for a view."""
    def decorator(f):
        from functools import wraps
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role not in roles:
                abort(403)
            return f(*args, **kwargs)
        return wrapped
    return decorator

# Ensure DB exists and default admin
with app.app_context():
    db.create_all()
    if not User.query.first():
        admin = User(username='admin', role='admin')
        admin.set_password('admin')
        db.session.add(admin)
        db.session.commit()

# ------------------ Routes ------------------
@app.route('/')
@login_required
def index():
    if current_user.role in ['admin', 'approver']:
        # Admin and approver see all PRs
        prs = PurchaseRequest.query.order_by(PurchaseRequest.created_at.desc()).limit(10).all()
    else:
        # Requester sees only their own PRs
        prs = PurchaseRequest.query.filter_by(created_by=current_user.id).order_by(
            PurchaseRequest.created_at.desc()).limit(10).all()

    return render_template('index.html', prs=prs)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        u = User.query.filter_by(username=username).first()
        if u and u.check_password(password):
            login_user(u)
            flash('Logged in successfully', 'success')
            next_page = request.args.get('next') or url_for('index')
            return redirect(next_page)
        flash('Invalid credentials', 'danger')
    return render_template('login.html')



@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out', 'info')
    return redirect(url_for('login'))

# ---- Suppliers CRUD ----
@app.route('/suppliers')
@login_required
def suppliers():
    if current_user.role != 'admin':  # Only admin can access
        flash('Access denied', 'danger')
        return redirect(url_for('index'))

    all_suppliers = Supplier.query.all()
    return render_template('suppliers.html', suppliers=all_suppliers)


@app.route('/suppliers/new', methods=['GET','POST'])
@login_required
@role_required('admin','approver')
def supplier_new():
    if request.method == 'POST':
        name = request.form['name']
        contact = request.form.get('contact')
        email = request.form.get('email')
        sup = Supplier(name=name, contact=contact, email=email)
        db.session.add(sup)
        db.session.commit()
        flash('Supplier created', 'success')
        return redirect(url_for('suppliers'))
    return render_template('supplier_form.html')

# ---- Purchase Requests ----
@app.route('/pr/new', methods=['GET', 'POST'])
@login_required
def pr_new():
    # Allow admin, approver, and requester to create PRs
    if current_user.role not in ['admin', 'approver', 'requester']:
        flash('You do not have permission to create a Purchase Request.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        pr = PurchaseRequest(
            title=title,
            description=description,
            created_by=current_user.id,
            status='draft'
        )
        db.session.add(pr)
        db.session.commit()

        # Handle line items
        item_names = request.form.getlist('item_name[]')
        quantities = request.form.getlist('quantity[]')
        units = request.form.getlist('unit[]')
        unit_prices = request.form.getlist('unit_price[]')

        for i in range(len(item_names)):
            line_item = LineItem(
                pr_id=pr.id,
                item_name=item_names[i],
                quantity=int(quantities[i]),
                unit=units[i],
                unit_price=float(unit_prices[i])
            )
            db.session.add(line_item)

        # Commit all line items
        db.session.commit()

        flash('Purchase Request created successfully!', 'success')
        return redirect(url_for('index'))

    # Render the existing PR form template
    return render_template('pr_form.html')

@app.route('/pr/<int:pr_id>')
@login_required
def pr_view(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)
    line_items = LineItem.query.filter_by(pr_id=pr.id).all()
    return render_template('pr_view.html', pr=pr, line_items=line_items)


@app.route('/pr/<int:pr_id>/approve', methods=['POST'])
@login_required
def pr_approve(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)

    if current_user.role not in ('approver', 'admin'):
        abort(403)

    action = request.form.get('action')
    comment = request.form.get('comment')

    if action == 'approve':
        pr.status = 'approved'
    elif action == 'reject':
        pr.status = 'rejected'

    log = ApprovalLog(
        pr_id=pr.id,
        user_id=current_user.id,
        action=action,
        comment=comment
    )
    db.session.add(log)
    db.session.commit()

    flash(f'PR {action}d successfully', 'success')
    return redirect(url_for('pr_view', pr_id=pr.id))


# ---- Simple API ----
@app.route('/api/prs')
def api_prs():
    prs = PurchaseRequest.query.all()
    data = []
    for p in prs:
        data.append({
            'id': p.id,
            'title': p.title,
            'status': p.status,
            'created_at': p.created_at.isoformat(),
            'total': p.total
        })
    return jsonify(data)

@app.route('/api/prs/<int:pr_id>')
def api_pr(pr_id):
    p = PurchaseRequest.query.get_or_404(pr_id)
    return jsonify({
        'id': p.id,
        'title': p.title,
        'description': p.description,
        'status': p.status,
        'created_at': p.created_at.isoformat(),
        'total': p.total
    })
# ------------------ User Management (Admin Only) ------------------

@app.route('/users')
@login_required
@role_required('admin')
def users():
    all_users = User.query.all()
    return render_template('users.html', users=all_users)


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def user_new():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']

        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'danger')
            return redirect(url_for('user_new'))

        new_user = User(username=username, role=role)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        flash('User created successfully', 'success')
        return redirect(url_for('users'))

    return render_template('user_form.html', user=None)


@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def user_edit(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        user.username = request.form['username']
        role = request.form['role']
        user.role = role

        new_password = request.form['password']
        if new_password:
            user.set_password(new_password)

        db.session.commit()
        flash('User updated successfully', 'success')
        return redirect(url_for('users'))

    return render_template('user_form.html', user=user)


@app.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def user_delete(user_id):
    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('users'))

    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully', 'info')
    return redirect(url_for('users'))

@app.route('/pr/<int:pr_id>/edit', methods=['GET', 'POST'])
@login_required
def pr_edit(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)

    # Optional: Restrict if only creator or admin/approver can edit
    if current_user.id != pr.created_by and current_user.role not in ('admin', 'approver'):
        abort(403)

    if request.method == 'POST':
        pr.title = request.form.get('title')
        pr.description = request.form.get('description')

        # Remove old items first
        LineItem.query.filter_by(pr_id=pr.id).delete()

        # Recreate from form
        item_names = request.form.getlist('item_name[]')
        quantities = request.form.getlist('quantity[]')
        units = request.form.getlist('unit[]')
        unit_prices = request.form.getlist('unit_price[]')

        total = 0
        for name, qty, price, unit_val in zip(item_names, quantities, unit_prices, units):
            if not name:
                continue
            try:
                qty = int(qty)
                price = float(price)
            except:
                qty = 1
                price = 0.0

            line_item = LineItem(
                pr_id=pr.id,
                item_name=name,
                quantity=qty,
                unit=unit_val,  # ✅ Add this
                unit_price=price

            )
            total += qty * price
            db.session.add(line_item)

        pr.total = total
        db.session.commit()
        update_balance_for_pr(pr.id)
        flash('Purchase Request updated', 'success')
        return redirect(url_for('pr_view', pr_id=pr.id))

    line_items = LineItem.query.filter_by(pr_id=pr.id).all()
    return render_template('pr_form_edit.html', pr=pr, line_items=line_items)

@app.route('/pr/<int:pr_id>/print')
@login_required
def pr_print(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)
    creator = User.query.get(pr.created_by)
    line_items = LineItem.query.filter_by(pr_id=pr.id).all()

    template_path = os.path.join('templates', 'pr_template.xlsx')
    output_path = os.path.join('generated', f'PR_{pr.id}.xlsx')

    os.makedirs('generated', exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb.active

    # Fill header fields
    combined_text = f"{pr.title} - {pr.description}" if pr.description else pr.title
    ws['B95'].value = combined_text
    ws['B99'] = creator.username if creator else 'Unknown'
    ws['E16'] = pr.created_at.strftime('%Y-%m-%d')

    # Line items
    start_row = 21
    last_lineitem_row = 94  # pre-set line item rows

    # First, clear all pre-set line item cells to avoid leftover data
    for r in range(start_row, last_lineitem_row + 1):
        for c in range(1, 7):  # Columns A-F
            ws.cell(row=r, column=c).value = None

    # Fill line items
    for i, item in enumerate(line_items):
        row = start_row + i
        ws[f'A{row}'] = i + 1
        ws[f'B{row}'] = item.item_name
        ws[f'C{row}'] = item.quantity
        ws[f'D{row}'] = item.unit
        ws[f'E{row}'] = item.unit_price
        ws[f'F{row}'] = item.quantity * item.unit_price

    # Place TOTAL immediately after last item
    total_row = start_row + len(line_items)
    ws[f'F{total_row}'] = pr.total

    # Delete any excess empty rows within pre-set line items
    if total_row < last_lineitem_row:
        ws.delete_rows(idx=total_row + 1, amount=last_lineitem_row - total_row)

    # Add fixed logo at B1
    image1_path = os.path.join('templates', 'Picture1.png')
    if os.path.exists(image1_path):
        from openpyxl.drawing.image import Image
        img1 = Image(image1_path)
        ws.add_image(img1, 'B1')

    wb.save(output_path)
    return send_file(output_path, as_attachment=True)


@app.route('/pr/<int:pr_id>/po', methods=['GET', 'POST'])
@login_required
def po_create(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)
    if pr.status != 'approved':
        flash('You can only create a Purchase Order for Approved PRs.', 'warning')
        return redirect(url_for('pr_view', pr_id=pr.id))

    # Get all line items for this PR
    items = LineItem.query.filter_by(pr_id=pr.id).all()

    # Get all suppliers from database
    suppliers = Supplier.query.all()

    if request.method == 'POST':
        supplier_names = request.form.getlist('supplier_name[]')
        brand_names = request.form.getlist('brand_name[]')
        quotation_prices = request.form.getlist('quotation_price[]')  # <-- NEW
        item_ids = request.form.getlist('item_id[]')

        for i, item_id in enumerate(item_ids):
            line_item = LineItem.query.get(int(item_id))
            if not line_item:
                continue  # skip if item not found

            po = PurchaseOrder(
                pr_id=pr.id,
                item_id=line_item.id,
                supplier_name=supplier_names[i],
                brand_name=brand_names[i],
                quotation_price=float(quotation_prices[i])  # ✅ Use input from form
            )
            db.session.add(po)

        db.session.commit()
        update_balance_for_pr(pr.id)
        flash('Purchase Order saved successfully!', 'success')
        return redirect(url_for('purchase_orders'))

    return render_template('po_form.html', pr=pr, items=items, suppliers=suppliers)




@app.route('/purchase_orders')
@login_required
def purchase_orders():
    if current_user.role == 'admin':
        prs = PurchaseRequest.query.join(PurchaseOrder).order_by(PurchaseRequest.created_at.desc()).all()
    else:
        # requester sees only POs linked to their PRs
        prs = PurchaseRequest.query.join(PurchaseOrder)\
              .filter(PurchaseRequest.created_by == current_user.id)\
              .order_by(PurchaseRequest.created_at.desc()).all()
    return render_template('po_list.html', prs=prs)


@app.route('/pr/<int:pr_id>/po/view')
@login_required
def po_view(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)
    orders = PurchaseOrder.query.filter_by(pr_id=pr_id).all()

    # Group by supplier
    supplier_groups = {}
    for o in orders:
        supplier_groups.setdefault(o.supplier_name, []).append(o)

    # Totals per supplier
    supplier_totals = {
        supplier: sum(o.quotation_price for o in items)
        for supplier, items in supplier_groups.items()
    }

    grand_total = sum(o.quotation_price for o in orders)

    return render_template(
        'po_view.html',
        pr=pr,
        supplier_groups=supplier_groups,
        supplier_totals=supplier_totals,
        grand_total=grand_total
    )

@app.route('/pr/<int:pr_id>/po_list')
@login_required
def po_list_by_pr(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)
    po_items = PurchaseOrder.query.filter_by(pr_id=pr.id).all()

    # Group by supplier
    supplier_groups = {}
    supplier_totals = {}
    grand_total = 0

    for po in po_items:
        supplier = po.supplier_name
        if supplier not in supplier_groups:
            supplier_groups[supplier] = []
            supplier_totals[supplier] = 0

        supplier_groups[supplier].append(po)

        # ✅ Total = quotation_price * quantity
        line_total = po.quotation_price * (po.item.quantity if po.item else 0)
        supplier_totals[supplier] += line_total
        grand_total += line_total

    return render_template(
        'po_list_by_pr.html',
        pr=pr,
        supplier_groups=supplier_groups,
        supplier_totals=supplier_totals,
        grand_total=grand_total
    )

@app.route('/po/update/<int:po_id>', methods=['POST'])
@login_required
def update_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    data = request.get_json()

    po.brand_name = data.get('brand_name', po.brand_name)
    po.quotation_price = data.get('quotation_price', po.quotation_price)

    db.session.commit()
    update_balance_for_pr(po.pr_id)
    return '', 200

@app.route('/po/delete/<int:po_id>', methods=['POST'])
@login_required
def delete_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    db.session.delete(po)
    db.session.commit()
    update_balance_for_pr(po.pr_id)
    flash('Purchase Order item deleted', 'success')
    return redirect(request.referrer or url_for('purchase_orders'))

# Admin: view all activities of a specific user
@app.route('/balance/<int:pr_id>')
@login_required
def balance_user(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)
    line_items = LineItem.query.filter_by(pr_id=pr_id).all()
    po_items = PurchaseOrder.query.filter_by(pr_id=pr_id).all()

    # Total PR
    pr_total = sum([li.subtotal for li in line_items])

    # Total PO = sum(quotation_price * quantity)
    po_total = sum([po.quotation_price * (po.item.quantity if po.item else 0) for po in po_items])

    balance = pr_total - po_total

    return render_template(
        'balance_by_user.html',
        pr=pr,
        line_items=line_items,
        po_items=po_items,
        pr_total=pr_total,
        po_total=po_total,
        balance=balance
    )
@app.route('/balance/user/<int:user_id>')
@login_required
def balance_by_user(user_id):
    # Fetch balances for all PRs created by this user
    prs = PurchaseRequest.query.filter_by(created_by=user_id).all()
    balances = []
    total_pr = total_po = total_balance = 0.0
    for pr in prs:
        update_balance_for_pr(pr.id)
        bal = Balance.query.filter_by(pr_id=pr.id).first()
        if bal:
            balances.append(bal)
            total_pr += bal.pr_total_amount or 0
            total_po += bal.po_total_amount or 0
            total_balance += bal.balance_amount or 0

    return render_template(
        'balance_by_user.html',
        balances=balances,
        total_pr=total_pr,
        total_po=total_po,
        total_balance=total_balance
    )


# Regular user: same totals included
@app.route('/balance')
@login_required
def balance_index():
    if current_user.role == 'admin':
        users = User.query.order_by(User.username).all()
        return render_template('balance_users.html', users=users)
    else:
        balances = []
        total_pr = total_po = total_balance = 0.0
        prs = PurchaseRequest.query.filter_by(created_by=current_user.id).all()
        for pr in prs:
            update_balance_for_pr(pr.id)
            bal = Balance.query.filter_by(pr_id=pr.id).first()
            if bal:
                balances.append(bal)
                total_pr += bal.pr_total_amount or 0
                total_po += bal.po_total_amount or 0
                total_balance += bal.balance_amount or 0

        return render_template(
            'balance_by_user.html',
            balances=balances,
            total_pr=total_pr,
            total_po=total_po,
            total_balance=total_balance
        )

@app.route('/balance/activity/<int:pr_id>')
@login_required
def balance_activity(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)

    # Permission check
    if current_user.role != 'admin' and current_user.id != pr.created_by:
        abort(403)

    # ✅ Fetch the creator
    creator = User.query.get(pr.created_by)

    # ✅ Ensure balances update
    update_balance_for_pr(pr_id)

    # ✅ Fetch lines & POs
    line_items = LineItem.query.filter_by(pr_id=pr_id).all()
    po_items = PurchaseOrder.query.filter_by(pr_id=pr_id).all()

    # ✅ DEBUG: Print values to confirm
    print("LINE ITEMS:", [(li.item_name, li.quantity, li.unit_price, li.subtotal) for li in line_items])
    print("PO ITEMS:", [(po.supplier_name, po.quotation_price) for po in po_items])

    return render_template(
        'balance_activity.html',
        pr=pr,
        creator=creator,
        line_items=line_items,
        po_items=po_items
    )

@app.route('/supplier/delete/<int:id>', methods=['POST'])
@login_required
def delete_supplier(id):
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('index'))
    # delete logic here

@app.route('/pr/submit/<int:pr_id>', methods=['POST'])
@login_required
def pr_submit(pr_id):
    pr = PurchaseRequest.query.get_or_404(pr_id)

    # Only the requester who created it can submit
    if pr.created_by != current_user.id or current_user.role != 'requester':
        flash('You do not have permission to submit this PR.', 'danger')
        return redirect(url_for('pr_view', pr_id=pr.id))

    pr.status = 'pending'
    db.session.commit()
    flash('Purchase Request submitted successfully!', 'success')
    return redirect(url_for('pr_view', pr_id=pr.id))

# Delete PR (Admin & Approver only)
@app.route('/pr/delete/<int:pr_id>', methods=['POST'])
@login_required
def pr_delete(pr_id):
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('index'))

    pr = PurchaseRequest.query.get_or_404(pr_id)

    # Check for related Purchase Orders
    if pr.purchase_orders:
        flash(
            'Cannot delete this PR because it has related Purchase Orders. '
            'Please delete the related POs first.', 'warning'
        )
        return redirect(url_for('index'))

    # Safe to delete PR
    db.session.delete(pr)
    db.session.commit()
    flash(f'PR #{pr_id} deleted successfully', 'success')
    return redirect(url_for('index'))

@app.route('/verify_admin_password', methods=['POST'])
@login_required
def verify_admin_password():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Access denied'})

    data = request.get_json()
    password = data.get('password')
    if current_user.check_password(password):
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Incorrect password'})

@app.route('/migrate_sqlite_to_postgres')
@login_required
def migrate_sqlite_to_postgres():
    if current_user.role != 'admin':
        return "Access denied. Only admin can perform migration.", 403

    sqlite_path = 'pms.db'
    postgres_url = app.config['SQLALCHEMY_DATABASE_URI']

    if not os.path.exists(sqlite_path):
        return "SQLite database (pms.db) not found in project directory.", 404

    try:
        # Create SQLAlchemy engines
        sqlite_engine = create_engine(f"sqlite:///{sqlite_path}")
        postgres_engine = create_engine(postgres_url)

        # Load SQLite data
        sqlite_meta = db.metadata
        sqlite_meta.reflect(bind=sqlite_engine)

        # Create PostgreSQL tables
        db.create_all()

        # Migrate table by table
        for table in sqlite_meta.tables.values():
            table_name = table.name
            rows = list(sqlite_engine.execute(table.select()))
            if rows:
                postgres_engine.execute(table.insert(), [dict(row) for row in rows])
                print(f"✅ Migrated {len(rows)} rows from {table_name}")

        return "✅ Migration completed successfully!", 200

    except Exception as e:
        print(e)
        return f"❌ Migration failed: {e}", 500

if __name__ == '__main__':
    # Use the PORT environment variable assigned by the hosting platform
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)